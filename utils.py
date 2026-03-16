from pathlib import Path
from typing import Tuple, Callable, List, Dict, Any
import soundfile as sf
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import welch, stft, istft, correlate, resample
from const import NPERSEG, TARGET_NORMALIZED_dB, SUFFICIENT_GRADE
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def load_audio(filepath, sr=None):
    signal, fs = sf.read(str(filepath), samplerate=sr)
    return signal, fs

def cut_audio(data, fs, t_start=False, t_end=False):
    start_idx = (int(t_start * fs) if t_start else 0)
    end_idx = (int(t_end * fs) if t_end else len(data)-1)
    return data[start_idx:end_idx], fs

def save_audio(filepath, data, sr=None):
    sf.write(filepath, data, samplerate=sr)

def compute_psd(data, fs, nperseg=NPERSEG):
    freqs, psd = welch(data, fs=fs, nperseg=nperseg)
    return freqs, psd

def compute_stft(data, fs, nperseg=NPERSEG, noverlap=None):
    return stft(data, fs, "hann", nperseg, noverlap, scaling="psd")

def compute_istft(Zxx, fs, nperseg=NPERSEG, noverlap=None):
    return istft(Zxx, fs, "hann", nperseg, noverlap)

def find_delay(ref, sig):
    corr = correlate(sig, ref, mode="full")
    lags = np.arange(-len(ref)+1, len(sig))
    delay = lags[np.argmax(corr)]
    return delay

def shift_signal(sig, delay):
    if delay > 0:
        return sig[delay:]
    elif delay < 0:
        return np.pad(sig, (abs(delay), 0))
    return sig

def align_to_ref(ref, sig):
    delay = find_delay(ref, sig)
    aligned = shift_signal(sig, delay)
    return aligned, delay

def align_length(ref, deg):
    L = min(len(ref), len(deg))
    return ref[:L], deg[:L]

def resample_audio(sig, orig_sr, target_sr):
    if orig_sr == target_sr:
        return sig
    duration = len(sig) / orig_sr
    target_len = int(duration * target_sr)
    return resample(sig, target_len)

def normalize(x):
    target_db = TARGET_NORMALIZED_dB
    rms = np.sqrt(np.mean(x**2))
    current_db = 20 * np.log10(rms + 1e-10)
    gain = 10 ** ((target_db - current_db) / 20)
    return x * gain

def plot_psd(freqs, psd):
    plt.figure()
    plt.plot(freqs, 10*np.log10(np.abs(psd)))
    plt.grid()
    plt.xlabel("Frequency (Hz)")
    plt.ylabel("Power Spectral Density")
    plt.show()

def plot_stft(f, t, Zxx):
    plt.figure()
    plt.pcolormesh(t, f, 10 * np.log10(np.abs(Zxx)), cmap="inferno", shading='gouraud')
    plt.xlabel('Time [s]')
    plt.ylabel('Frequency [Hz]')
    plt.title('STFT Magnitude')
    plt.colorbar(label='Amplitude')
    plt.show()

def metric_color(metric, value):
    if metric == "GCC_delay":
        score = max(0, 1 - abs(value) / 0.05)
    else:
        thresholds = SUFFICIENT_GRADE
        score = max(0, min(1, value / thresholds.get(metric, 1)))
    r = int(255 * (1 - score))
    g = int(255 * score)
    return f"{r:02X}{g:02X}00"

def color_excel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for col_idx, name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if name == "Test":
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            try:
                val = float(cell.value)
            except:
                continue
            if name.startswith("Δ"):
                if val > 0:
                    color = "99FF99"
                elif val < 0:
                    color = "FF9999"
                else:
                    continue
            else:
                color = metric_color(name, val)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    wb.save(filename)

def append_test_to_excel(test_name, metrics, filename=r"C:\Users\User\Desktop\sound\wav\nr\nr_results.xlsx"):
    row = {"Test": test_name, **metrics}
    try:
        df = pd.read_excel(filename)
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    except FileNotFoundError:
        df = pd.DataFrame([row])
    if len(df) > 0:
        ref = df.iloc[0]
        for k in metrics.keys():
            delta_col = f"Δ{k}"
            if delta_col not in df.columns:
                df[delta_col] = None
            if k == "GCC_delay":
                df.loc[df.index[-1], delta_col] = abs(ref[k]) - abs(df.loc[df.index[-1], k])
            else:
                df.loc[df.index[-1], delta_col] = df.loc[df.index[-1], k] - ref[k]
    df.to_excel(filename, index=False)
    color_excel(filename)

def run_algos(sig, fs, algos: List[Callable]):
    name_parts = []
    result = sig
    for algo in algos:
        result, params = algo(result)
        name = getattr(algo, "__name__", algo.func.__name__)
        params.pop("sr")
        name_parts.append(f"{name}: [{', '.join(f'{k}={v}' for k, v in params.items())}]")
    return ", ".join(name_parts), result

def return_params(func):
    def inner(*args, **kwargs):
        return func(*args, **kwargs), kwargs
    inner.__name__ = func.__name__
    return inner
